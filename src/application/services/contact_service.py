import os
import re
from itertools import chain
from typing import Optional

from src.domain.entities import Contact
from src.domain.enums import ContactImportSource
from src.domain.interfaces import ContactRepository


class ContactService:
    PHONE_HEADER_NAMES = {
        "phone", "mobile", "telephone", "tel",
        "رقم", "الهاتف", "رقم الهاتف", "الجوال", "رقم الجوال", "موبايل",
    }
    NAME_HEADER_NAMES = {"name", "full name", "contact", "الاسم", "اسم", "اسم العميل"}
    GROUP_HEADER_NAMES = {"group", "group name", "المجموعة", "مجموعة"}
    NOTES_HEADER_NAMES = {"notes", "note", "ملاحظات", "ملاحظة"}

    def __init__(self, repo: ContactRepository):
        self._repo = repo

    def add(self, contact: Contact) -> Contact:
        return self._repo.add(contact)

    def update(self, contact: Contact) -> bool:
        return self._repo.update(contact)

    def delete(self, contact_id: int) -> bool:
        return self._repo.delete(contact_id)

    def delete_all(self) -> int:
        return self._repo.delete_all()

    def delete_group(self, group_name: str) -> int:
        return self._repo.delete_group(group_name)

    def get_all(self, group: Optional[str] = None) -> list[Contact]:
        return self._repo.get_all(group)

    def get_by_id(self, contact_id: int) -> Optional[Contact]:
        return self._repo.get_by_id(contact_id)

    def get_groups(self) -> list[str]:
        return self._repo.get_groups()

    def count(self, group: Optional[str] = None) -> int:
        return self._repo.count(group)

    def search(self, q: str) -> list[Contact]:
        return self._repo.search(q)

    @staticmethod
    def normalize_import_phone(raw_phone) -> str:
        """Normalize imported phone numbers and restore a missing local leading zero.

        Spreadsheet applications often convert 079xxxxxxx to a number and store it as
        79xxxxxxx. This method deliberately restores the leading 0. Jordan numbers
        written with +962 / 00962 are also converted to their local 0-prefixed form.
        """
        if raw_phone is None:
            return ""

        if isinstance(raw_phone, float) and raw_phone.is_integer():
            raw_phone = int(raw_phone)

        raw_text = str(raw_phone).strip()
        if not raw_text:
            return ""

        digits = re.sub(r"[^0-9]", "", raw_text)
        if not digits:
            return ""

        if digits.startswith("00962") and len(digits) > 5:
            local = digits[5:].lstrip("0")
            return f"0{local}" if local else ""
        if digits.startswith("962") and len(digits) > 3:
            local = digits[3:].lstrip("0")
            return f"0{local}" if local else ""

        if not digits.startswith("0"):
            digits = "0" + digits
        return digits

    @staticmethod
    def _cell_text(value) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        return str(value).strip()

    @staticmethod
    def _header_text(value) -> str:
        return re.sub(r"\s+", " ", ContactService._cell_text(value).lower()).strip()

    def _existing_phone_keys(self) -> set[str]:
        keys = set()
        for contact in self._repo.get_all():
            phone = self.normalize_import_phone(contact.phone)
            if phone:
                keys.add(phone)
        return keys

    def _append_import_contact(
        self,
        contacts: list[Contact],
        errors: list[str],
        seen_phones: set[str],
        raw_phone,
        name: str,
        group_name: str,
        notes: str,
        source: ContactImportSource,
        row_label: str,
        unnamed_counter: list[int],
    ) -> None:
        phone = self.normalize_import_phone(raw_phone)
        if not phone:
            errors.append(f"{row_label}: رقم هاتف فارغ أو غير صالح ({raw_phone})")
            return

        if phone in seen_phones:
            errors.append(f"{row_label}: تم تجاهل رقم مكرر ({phone})")
            return

        clean_name = self._cell_text(name)
        if not clean_name:
            unnamed_counter[0] += 1
            clean_name = f"رقم {unnamed_counter[0]}"

        contact = Contact(
            name=clean_name,
            phone=phone,
            group_name=self._cell_text(group_name) or "عام",
            notes=self._cell_text(notes),
            import_source=source,
        )
        validation_errors = contact.validate()
        if validation_errors:
            errors.append(
                f"{row_label}: {', '.join(validation_errors)} (رقم: {phone})"
            )
            return

        seen_phones.add(phone)
        contacts.append(contact)

    def import_from_txt(
        self, file_path: str, default_group: str = "عام"
    ) -> tuple[int, list[str]]:
        contacts: list[Contact] = []
        errors: list[str] = []
        unnamed_counter = [0]
        seen_phones = self._existing_phone_keys()
        source = (
            ContactImportSource.CSV
            if os.path.splitext(file_path)[1].lower() == ".csv"
            else ContactImportSource.TXT
        )

        with open(file_path, encoding="utf-8-sig") as file:
            for line_no, raw_line in enumerate(file, 1):
                line = raw_line.strip()
                if not line or line.startswith("#"):
                    continue

                if "," in line:
                    parts = [part.strip() for part in line.split(",", 1)]
                    raw_phone = parts[0]
                    name = parts[1]
                elif "\t" in line:
                    parts = [part.strip() for part in line.split("\t", 1)]
                    raw_phone = parts[0]
                    name = parts[1]
                else:
                    raw_phone = line
                    name = ""

                if line_no == 1 and self._header_text(raw_phone) in self.PHONE_HEADER_NAMES:
                    continue

                self._append_import_contact(
                    contacts=contacts,
                    errors=errors,
                    seen_phones=seen_phones,
                    raw_phone=raw_phone,
                    name=name,
                    group_name=default_group,
                    notes="",
                    source=source,
                    row_label=f"سطر {line_no}",
                    unnamed_counter=unnamed_counter,
                )

        if contacts:
            self._repo.bulk_add(contacts)
        return len(contacts), errors

    def import_from_excel(
        self, file_path: str, default_group: str = "عام"
    ) -> tuple[int, list[str]]:
        from openpyxl import load_workbook

        contacts: list[Contact] = []
        errors: list[str] = []
        unnamed_counter = [0]
        seen_phones = self._existing_phone_keys()

        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            first_row = next(rows, None)
            if first_row is None:
                return 0, ["ملف Excel فارغ"]

            headers = [self._header_text(value) for value in first_row]

            def find_column(names: set[str]):
                for index, header in enumerate(headers):
                    if header in names:
                        return index
                return None

            phone_col = find_column(self.PHONE_HEADER_NAMES)
            name_col = find_column(self.NAME_HEADER_NAMES)
            group_col = find_column(self.GROUP_HEADER_NAMES)
            notes_col = find_column(self.NOTES_HEADER_NAMES)
            has_header = phone_col is not None

            if not has_header:
                phone_col = 0
                name_col = 1 if len(first_row) > 1 else None
                data_rows = chain(((1, first_row),), enumerate(rows, start=2))
            else:
                data_rows = enumerate(rows, start=2)

            for row_no, row in data_rows:
                if not row or all(
                    value is None or self._cell_text(value) == "" for value in row
                ):
                    continue

                raw_phone = (
                    row[phone_col]
                    if phone_col is not None and phone_col < len(row)
                    else ""
                )
                name = (
                    row[name_col]
                    if name_col is not None and name_col < len(row)
                    else ""
                )
                group = (
                    row[group_col]
                    if group_col is not None and group_col < len(row)
                    else default_group
                )
                notes = (
                    row[notes_col]
                    if notes_col is not None and notes_col < len(row)
                    else ""
                )

                self._append_import_contact(
                    contacts=contacts,
                    errors=errors,
                    seen_phones=seen_phones,
                    raw_phone=raw_phone,
                    name=self._cell_text(name),
                    group_name=self._cell_text(group) or default_group,
                    notes=self._cell_text(notes),
                    source=ContactImportSource.CSV,
                    row_label=f"صف {row_no}",
                    unnamed_counter=unnamed_counter,
                )
        finally:
            workbook.close()

        if contacts:
            self._repo.bulk_add(contacts)
        return len(contacts), errors
